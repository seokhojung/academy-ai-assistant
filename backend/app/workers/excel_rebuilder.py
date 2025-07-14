import os
import tempfile
import shutil
from datetime import datetime
from typing import Dict, Any, Optional
import portalocker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from celery import current_task
from google.cloud import storage
from app.workers.celery_app import celery_app
from app.core.config import settings
from app.services.student_service import StudentService
from app.services.teacher_service import TeacherService
from app.services.material_service import MaterialService
from app.core.database import get_session


class ExcelRebuilder:
    """Excel 파일 자동 재생성 클래스"""
    
    def __init__(self):
        self.gcs_client = None
        self._init_gcs_client()
    
    def _init_gcs_client(self):
        """GCS 클라이언트 초기화"""
        try:
            if os.path.exists(settings.gcs_credentials_path):
                self.gcs_client = storage.Client.from_service_account_json(
                    settings.gcs_credentials_path
                )
            else:
                # 환경 변수에서 인증 정보 사용
                self.gcs_client = storage.Client()
        except Exception as e:
            print(f"GCS 클라이언트 초기화 실패: {e}")
            self.gcs_client = None
    
    def create_student_excel(self, data: Dict[str, Any]) -> str:
        """학생 데이터로 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "학생 목록"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 설정
        headers = ["ID", "이름", "이메일", "전화번호", "학년", "수강료", "납부일", "상태", "등록일"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 입력
        for row, student in enumerate(data.get("students", []), 2):
            ws.cell(row=row, column=1, value=student.get("id"))
            ws.cell(row=row, column=2, value=student.get("name"))
            ws.cell(row=row, column=3, value=student.get("email"))
            ws.cell(row=row, column=4, value=student.get("phone"))
            ws.cell(row=row, column=5, value=student.get("grade"))
            ws.cell(row=row, column=6, value=student.get("tuition_fee"))
            ws.cell(row=row, column=7, value=student.get("tuition_due_date"))
            ws.cell(row=row, column=8, value="활성" if student.get("is_active") else "비활성")
            ws.cell(row=row, column=9, value=student.get("created_at"))
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def create_teacher_excel(self, data: Dict[str, Any]) -> str:
        """강사 데이터로 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "강사 목록"
        
        # 헤더 설정
        headers = ["ID", "이름", "이메일", "전화번호", "전문분야", "상태", "등록일"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터 입력
        for row, teacher in enumerate(data.get("teachers", []), 2):
            ws.cell(row=row, column=1, value=teacher.get("id"))
            ws.cell(row=row, column=2, value=teacher.get("name"))
            ws.cell(row=row, column=3, value=teacher.get("email"))
            ws.cell(row=row, column=4, value=teacher.get("phone"))
            ws.cell(row=row, column=5, value=teacher.get("specialty"))
            ws.cell(row=row, column=6, value="활성" if teacher.get("is_active") else "비활성")
            ws.cell(row=row, column=7, value=teacher.get("created_at"))
        
        return wb
    
    def create_material_excel(self, data: Dict[str, Any]) -> str:
        """교재 데이터로 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "교재 목록"
        
        # 헤더 설정
        headers = ["ID", "제목", "저자", "출판사", "ISBN", "재고", "가격", "상태", "등록일"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터 입력
        for row, material in enumerate(data.get("materials", []), 2):
            ws.cell(row=row, column=1, value=material.get("id"))
            ws.cell(row=row, column=2, value=material.get("title"))
            ws.cell(row=row, column=3, value=material.get("author"))
            ws.cell(row=row, column=4, value=material.get("publisher"))
            ws.cell(row=row, column=5, value=material.get("isbn"))
            ws.cell(row=row, column=6, value=material.get("stock"))
            ws.cell(row=row, column=7, value=material.get("price"))
            ws.cell(row=row, column=8, value="활성" if material.get("is_active") else "비활성")
            ws.cell(row=row, column=9, value=material.get("created_at"))
        
        return wb
    
    def save_excel_with_lock(self, wb: Workbook, file_path: str) -> bool:
        """portalocker를 사용하여 파일 잠금 후 Excel 저장"""
        try:
            # 임시 파일 생성
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_path = temp_file.name
            temp_file.close()
            
            # 임시 파일에 저장
            wb.save(temp_path)
            
            # portalocker로 파일 잠금 후 원본 파일로 이동
            with portalocker.Lock(file_path, timeout=30) as lock:
                shutil.move(temp_path, file_path)
                lock.release()
            
            return True
        except Exception as e:
            print(f"Excel 파일 저장 실패: {e}")
            # 임시 파일 정리
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            return False
    
    def upload_to_gcs(self, file_path: str, gcs_path: str) -> Optional[str]:
        """GCS에 파일 업로드"""
        if not self.gcs_client:
            return None
        
        try:
            bucket = self.gcs_client.bucket(settings.gcs_bucket_name)
            blob = bucket.blob(gcs_path)
            
            # 파일 업로드
            blob.upload_from_filename(file_path)
            
            # 공개 URL 반환
            return blob.public_url
        except Exception as e:
            print(f"GCS 업로드 실패: {e}")
            return None


@celery_app.task(bind=True, max_retries=3)
def rebuild_student_excel(self, file_type: str = "students") -> Dict[str, Any]:
    """학생 Excel 파일 재생성 태스크"""
    try:
        # 진행 상황 업데이트
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "데이터 수집 중..."}
        )
        
        # 데이터베이스에서 학생 데이터 수집
        with next(get_session()) as session:
            student_service = StudentService(session)
            students = student_service.get_students(limit=1000)
            
            # 데이터 변환
            student_data = []
            for student in students:
                student_data.append({
                    "id": student.id,
                    "name": student.name,
                    "email": student.email,
                    "phone": student.phone,
                    "grade": student.grade,
                    "tuition_fee": student.tuition_fee,
                    "tuition_due_date": student.tuition_due_date.isoformat() if student.tuition_due_date else None,
                    "is_active": student.is_active,
                    "created_at": student.created_at.isoformat()
                })
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 30, "total": 100, "status": "Excel 파일 생성 중..."}
        )
        
        # Excel 파일 생성
        rebuilder = ExcelRebuilder()
        wb = rebuilder.create_student_excel({"students": student_data})
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 60, "total": 100, "status": "파일 저장 중..."}
        )
        
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        local_file_path = f"temp/students_{timestamp}.xlsx"
        gcs_file_path = f"excel/students_{timestamp}.xlsx"
        
        # temp 디렉토리 생성
        os.makedirs("temp", exist_ok=True)
        
        # 파일 저장 (잠금 포함)
        if not rebuilder.save_excel_with_lock(wb, local_file_path):
            raise Exception("Excel 파일 저장 실패")
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 80, "total": 100, "status": "GCS 업로드 중..."}
        )
        
        # GCS 업로드
        gcs_url = rebuilder.upload_to_gcs(local_file_path, gcs_file_path)
        
        # 임시 파일 정리
        if os.path.exists(local_file_path):
            os.unlink(local_file_path)
        
        current_task.update_state(
            state="SUCCESS",
            meta={"current": 100, "total": 100, "status": "완료"}
        )
        
        return {
            "success": True,
            "file_path": gcs_file_path,
            "gcs_url": gcs_url,
            "student_count": len(student_data),
            "timestamp": timestamp
        }
        
    except Exception as e:
        # 재시도 로직
        if self.request.retries < self.max_retries:
            raise self.retry(countdown=60 * (self.request.retries + 1))
        else:
            return {
                "success": False,
                "error": str(e),
                "retries": self.request.retries
            }


@celery_app.task(bind=True, max_retries=3)
def rebuild_teacher_excel(self, file_type: str = "teachers") -> Dict[str, Any]:
    """강사 Excel 파일 재생성 태스크"""
    try:
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "강사 데이터 수집 중..."}
        )
        
        # 데이터베이스에서 강사 데이터 수집
        with next(get_session()) as session:
            teacher_service = TeacherService(session)
            teachers = teacher_service.get_teachers(limit=1000)
            
            # 데이터 변환
            teacher_data = []
            for teacher in teachers:
                teacher_data.append({
                    "id": teacher.id,
                    "name": teacher.name,
                    "email": teacher.email,
                    "phone": teacher.phone,
                    "specialty": teacher.specialty,
                    "is_active": teacher.is_active,
                    "created_at": teacher.created_at.isoformat()
                })
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 30, "total": 100, "status": "Excel 파일 생성 중..."}
        )
        
        # Excel 파일 생성
        rebuilder = ExcelRebuilder()
        wb = rebuilder.create_teacher_excel({"teachers": teacher_data})
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 60, "total": 100, "status": "파일 저장 중..."}
        )
        
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        local_file_path = f"temp/teachers_{timestamp}.xlsx"
        gcs_file_path = f"excel/teachers_{timestamp}.xlsx"
        
        os.makedirs("temp", exist_ok=True)
        
        if not rebuilder.save_excel_with_lock(wb, local_file_path):
            raise Exception("Excel 파일 저장 실패")
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 80, "total": 100, "status": "GCS 업로드 중..."}
        )
        
        # GCS 업로드
        gcs_url = rebuilder.upload_to_gcs(local_file_path, gcs_file_path)
        
        # 임시 파일 정리
        if os.path.exists(local_file_path):
            os.unlink(local_file_path)
        
        current_task.update_state(
            state="SUCCESS",
            meta={"current": 100, "total": 100, "status": "완료"}
        )
        
        return {
            "success": True,
            "file_path": gcs_file_path,
            "gcs_url": gcs_url,
            "teacher_count": len(teacher_data),
            "timestamp": timestamp
        }
        
    except Exception as e:
        if self.request.retries < self.max_retries:
            raise self.retry(countdown=60 * (self.request.retries + 1))
        else:
            return {
                "success": False,
                "error": str(e),
                "retries": self.request.retries
            }


@celery_app.task(bind=True, max_retries=3)
def rebuild_material_excel(self, file_type: str = "materials") -> Dict[str, Any]:
    """교재 Excel 파일 재생성 태스크"""
    try:
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 0, "total": 100, "status": "교재 데이터 수집 중..."}
        )
        
        # 데이터베이스에서 교재 데이터 수집
        with next(get_session()) as session:
            material_service = MaterialService(session)
            materials = material_service.get_materials(limit=1000)
            
            # 데이터 변환
            material_data = []
            for material in materials:
                material_data.append({
                    "id": material.id,
                    "title": material.title,
                    "author": material.author,
                    "publisher": material.publisher,
                    "isbn": material.isbn,
                    "stock": material.stock,
                    "price": material.price,
                    "is_active": material.is_active,
                    "created_at": material.created_at.isoformat()
                })
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 30, "total": 100, "status": "Excel 파일 생성 중..."}
        )
        
        # Excel 파일 생성
        rebuilder = ExcelRebuilder()
        wb = rebuilder.create_material_excel({"materials": material_data})
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 60, "total": 100, "status": "파일 저장 중..."}
        )
        
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        local_file_path = f"temp/materials_{timestamp}.xlsx"
        gcs_file_path = f"excel/materials_{timestamp}.xlsx"
        
        os.makedirs("temp", exist_ok=True)
        
        if not rebuilder.save_excel_with_lock(wb, local_file_path):
            raise Exception("Excel 파일 저장 실패")
        
        current_task.update_state(
            state="PROGRESS",
            meta={"current": 80, "total": 100, "status": "GCS 업로드 중..."}
        )
        
        # GCS 업로드
        gcs_url = rebuilder.upload_to_gcs(local_file_path, gcs_file_path)
        
        # 임시 파일 정리
        if os.path.exists(local_file_path):
            os.unlink(local_file_path)
        
        current_task.update_state(
            state="SUCCESS",
            meta={"current": 100, "total": 100, "status": "완료"}
        )
        
        return {
            "success": True,
            "file_path": gcs_file_path,
            "gcs_url": gcs_url,
            "material_count": len(material_data),
            "timestamp": timestamp
        }
        
    except Exception as e:
        if self.request.retries < self.max_retries:
            raise self.retry(countdown=60 * (self.request.retries + 1))
        else:
            return {
                "success": False,
                "error": str(e),
                "retries": self.request.retries
            } 